from openpyxl import *


def get():
    wb = load_workbook('program音色表.xlsx')

    ws = wb.active
    # 所有行
    keys = []
    values = []
    for row in ws.iter_rows():
        for cell in row:
            # print(cell.value)
            try:
                keys.append(int(cell.value))
            except ValueError:
                values.append(cell.value)
    # # 所有列
    # for column in ws.iter_cols():
    #     for cell in column:
    #         print(cell.value)
    out = ""
    index = 0
    for i in keys:
        out += ", \"" + str(i) + "\": \"" + values[index] + "\""
        index += 1

    print(out)
